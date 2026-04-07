#!/usr/bin/env python3
from __future__ import annotations

"""
Export Genesis OS Phase 1 workbook sheets into machine-readable seed files.

The exporter normalizes non-canonical sheet headers (for example
"Inside Chemistry Weight" -> "inside_chemistry_weight") so the
resulting CSV/JSON files can load cleanly into code and databases.

This script exports:
- workbook-backed import tables as CSV and JSON
- polarity_pair_engine_v10 as structured JSON
- workbook_import_registry seed JSON
- workbook_summary metadata
"""

import argparse
import hashlib
import json
import re
from pathlib import Path
from datetime import datetime, timezone

import pandas as pd
import openpyxl

def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()

def normalize_col(name: object) -> str:
    s = str(name).strip().lower()
    s = s.replace("–", "-").replace("—", "-").replace("%", "pct")
    s = re.sub(r"[()]+", " ", s)
    s = re.sub(r"[/]+", " ", s)
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s

def normalized_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [normalize_col(c) for c in out.columns]
    return out

def clean_matrix(df: pd.DataFrame, name_col: str) -> pd.DataFrame:
    cols = list(df.iloc[0])
    cols[0] = name_col
    clean = df.iloc[1:].copy()
    clean.columns = cols
    clean.reset_index(drop=True, inplace=True)
    return normalized_df(clean)

def export() -> None:
    parser = argparse.ArgumentParser(description="Export workbook-backed Phase 1 seed data.")
    parser.add_argument("--workbook", required=True, help="Path to genesis_os_final_quantum_scoring_workbook_v10_polarity.xlsx")
    parser.add_argument("--output-dir", required=True, help="Output directory")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    out_dir = Path(args.output_dir)
    csv_dir = out_dir / "csv"
    json_dir = out_dir / "json"
    csv_dir.mkdir(parents=True, exist_ok=True)
    json_dir.mkdir(parents=True, exist_ok=True)

    xls = pd.ExcelFile(workbook_path, engine="openpyxl")
    wb = openpyxl.load_workbook(workbook_path, data_only=False, read_only=True)

    answer_bank = normalized_df(pd.read_excel(xls, "AnswerBank"))
    answer_pair = normalized_df(pd.read_excel(xls, "AnswerPairQuantum"))

    questions = (
        answer_bank
        .groupby(["question_number_int","phase","dimension_primary","dimension_secondary","question_weight_v3"], as_index=False)
        .agg(answer_count=("answer_id","count"),
             available_letters=("answer_letter", lambda s: "".join(s.astype(str).tolist())),
             answer_ids=("answer_id", lambda s: "|".join(s.astype(str).tolist())))
    )
    questions["question_id"] = questions["question_number_int"].apply(lambda q: f"Q{int(q):03d}")
    questions = questions[["question_id","question_number_int","phase","question_weight_v3","dimension_primary","dimension_secondary","answer_count","available_letters","answer_ids"]]

    datasets = {
        "questions": questions,
        "answer_bank": answer_bank,
        "answer_pair_quantum": answer_pair,
        "archetype_matrix": clean_matrix(pd.read_excel(xls, "ArchetypeMatrix", header=None), "row_label"),
        "shadow_matrix": clean_matrix(pd.read_excel(xls, "ShadowMatrix", header=None), "row_label"),
        "zodiac_lookup": normalized_df(pd.read_excel(xls, "ZodiacLookup")),
        "numerology_lookup": normalized_df(pd.read_excel(xls, "NumerologyLookup")),
        "schema_weights": normalized_df(pd.read_excel(xls, "Schema")),
        "feedback_data_fields": normalized_df(pd.read_excel(xls, "FeedbackData")),
        "polarity_axes_v10": normalized_df(pd.read_excel(xls, "Polarity_Axes_v10")),
        "polarity_role_seeds_v10": normalized_df(pd.read_excel(xls, "Polarity_RoleSeeds_v10")),
        "polarity_question_map_v10": normalized_df(pd.read_excel(xls, "Polarity_QuestionMap_v10")),
        "polarity_answer_bank_v10": normalized_df(pd.read_excel(xls, "Polarity_AnswerBank_v10")),
        "rewrite_rules_v10": normalized_df(pd.read_excel(xls, "Rewrite_Rules_v10")),
    }

    manifest = []
    for name, df in datasets.items():
        csv_path = csv_dir / f"{name}.csv"
        json_path = json_dir / f"{name}.json"
        df.to_csv(csv_path, index=False)
        df.to_json(json_path, orient="records", force_ascii=False, indent=2)
        manifest.append({
            "name": name,
            "rows": int(len(df)),
            "columns": list(df.columns),
            "csv": str(csv_path),
            "json": str(json_path),
            "sha256_csv": sha256_file(csv_path),
            "sha256_json": sha256_file(json_path),
        })

    ws = wb["Polarity_PairEngine_v10"]
    axes = []
    for r in range(4, 14):
        axes.append({
            "user_a_axis": ws.cell(r, 1).value,
            "user_a_example_value": ws.cell(r, 2).value,
            "user_b_axis": ws.cell(r, 4).value,
            "user_b_example_value": ws.cell(r, 5).value,
            "dimension": ws.cell(r, 7).value,
            "fit_formula_type": ws.cell(r, 8).value,
            "fit_formula_excel": ws.cell(r, 9).value,
        })

    pair_engine = {
        "title": ws["A1"].value,
        "axes": axes,
        "aggregate_formulas": {
            "polarity_dimension_score_0_100": ws["I15"].value,
            "attraction_pattern_norm": ws["I16"].value,
            "chemistry_layer_contribution": ws["I17"].value,
        },
        "notes": [
            ws["A16"].value,
            ws["G18"].value,
            ws["G19"].value,
        ],
    }
    pair_engine_path = json_dir / "polarity_pair_engine_v10.json"
    pair_engine_path.write_text(json.dumps(pair_engine, indent=2), encoding="utf-8")
    manifest.append({
        "name": "polarity_pair_engine_v10",
        "rows": len(axes),
        "columns": list(axes[0].keys()),
        "json": str(pair_engine_path),
        "sha256_json": sha256_file(pair_engine_path),
    })

    imported_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    checksum = sha256_file(workbook_path)
    workbook_import_registry = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        workbook_import_registry.append({
            "import_id": f"{workbook_path.stem}:{sheet_name}:{checksum[:12]}",
            "artifact_type": "xlsx_sheet",
            "workbook_version": workbook_path.stem,
            "sheet_name": sheet_name,
            "sheet_rows": ws.max_row,
            "data_rows": max(ws.max_row - 1, 0),
            "source_checksum": checksum,
            "imported_at": imported_at,
            "import_job_id": f"phase1_export:{checksum[:12]}",
            "score_contract_version": "phase1.v1",
        })
    registry_path = json_dir / "workbook_import_registry.seed.json"
    registry_path.write_text(json.dumps(workbook_import_registry, indent=2), encoding="utf-8")

    summary = {
        "workbook_file": workbook_path.name,
        "workbook_sha256": checksum,
        "generated_at_utc": imported_at,
        "sheet_count": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
        "datasets": manifest,
        "workbook_import_registry": str(registry_path),
    }
    (json_dir / "workbook_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")

if __name__ == "__main__":
    export()
